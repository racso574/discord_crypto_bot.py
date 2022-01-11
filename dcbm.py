import os
import discord
from dotenv import load_dotenv
from discord.ext import tasks, commands
from openpyxl import load_workbook
import urllib.request
import re
import time
from datetime import date




load_dotenv()
TOKEN = os.getenv('DISCORD_TOKEN')
client = discord.Client()
client = commands.Bot(command_prefix='')
client.remove_command('help')




@client.event
async def on_ready():
    print(f'{client.user} uwu')
    cl.start()
    mamial.start()

@client.command()
async def cp(ctx):
    webdatabtc = urllib.request.urlopen('https://coinmarketcap.com/currencies' + os.getenv('c1'))
    cdatabtc = re.findall(r"<span>\$(\S{9})", webdatabtc.read().decode())
    webdataeth = urllib.request.urlopen('https://coinmarketcap.com/currencies' + os.getenv('c2'))
    cdataeth = re.findall(r"<span>\$(\S{8})", webdataeth.read().decode())
    webdatabnb = urllib.request.urlopen('https://coinmarketcap.com/currencies' + os.getenv('c3'))
    cdatabnb = re.findall(r"<span>\$(\S{6})", webdatabnb.read().decode())
    webdatasol = urllib.request.urlopen('https://coinmarketcap.com/currencies' + os.getenv('c4'))
    cdatasol = re.findall(r"<span>\$(\S{6})", webdatasol.read().decode())
    webdataada = urllib.request.urlopen('https://coinmarketcap.com/currencies' + os.getenv('c5'))
    cdataada = re.findall(r"<span>\$(\S{4})", webdataada.read().decode())
    embed = discord.Embed(title="", description=f"")
    embed.add_field(name="BTC", value=cdatabtc[0], inline=False)
    embed.add_field(name="ETH", value=cdataeth[0], inline=False)
    embed.add_field(name="BNB", value=cdatabnb[0], inline=False)
    embed.add_field(name="SOL", value=cdatasol[0], inline=False)
    embed.add_field(name="ADA", value=cdataada[0], inline=False)
    await ctx.send(embed=embed)



@tasks.loop(seconds=60)
async def cl():
    webdatabtc = urllib.request.urlopen('https://coinmarketcap.com/currencies' + os.getenv('c1'))
    cdatabtc = re.findall(r"<span>\$(\S{9})", webdatabtc.read().decode())
    webdataeth = urllib.request.urlopen('https://coinmarketcap.com/currencies' + os.getenv('c2'))
    cdataeth = re.findall(r"<span>\$(\S{8})", webdataeth.read().decode())
    webdatabnb = urllib.request.urlopen('https://coinmarketcap.com/currencies' + os.getenv('c3'))
    cdatabnb = re.findall(r"<span>\$(\S{6})", webdatabnb.read().decode())
    webdatasol = urllib.request.urlopen('https://coinmarketcap.com/currencies' + os.getenv('c4'))
    cdatasol = re.findall(r"<span>\$(\S{6})", webdatasol.read().decode())
    webdataada = urllib.request.urlopen('https://coinmarketcap.com/currencies' + os.getenv('c5'))
    cdataada = re.findall(r"<span>\$(\S{4})", webdataada.read().decode())
    embed = discord.Embed(title="Cryptocurrency Values", description=f"")
    embed.add_field(name="BTC", value=cdatabtc[0], inline=False)
    embed.add_field(name="ETH", value=cdataeth[0], inline=False)
    embed.add_field(name="BNB", value=cdatabnb[0], inline=False)
    embed.add_field(name="SOL", value=cdatasol[0], inline=False)
    embed.add_field(name="ADA", value=cdataada[0], inline=False)
    channel = client.get_channel(id=927667259495829555)
    myid = '<@' + os.getenv('uid') + '>'
    if time.strftime("%H%M") == os.getenv('h1'):
        await channel.send(myid)
        await channel.send(embed=embed)
    elif time.strftime("%H%M") == os.getenv('h2'):
        await channel.send(myid)
        await channel.send(embed=embed)
    elif time.strftime("%H%M") == os.getenv('h3'):
        await channel.send(myid)
        await channel.send(embed=embed)

#allowed_mentions = discord.AllowedMentions.all()
#content="@everyone",allowed_mentions=allowed_mentions,

@tasks.loop(seconds=60)
async def mamial():
    webdata1 = urllib.request.urlopen('https://coinmarketcap.com/currencies' + os.getenv('c1'))
    cdata1 = re.findall(r"<span>\$(\S{2})", webdata1.read().decode())
    embed = discord.Embed(title="", description=f"")
    embed.add_field(name="BTC", value=cdata1[0], inline=False)
    channel = client.get_channel(id=927667259495829555)
    myid = '<@' + os.getenv('uid') + '>'
    xl = load_workbook('db.xlsx')
    edit = xl['Sheet1']
    b1 = str(edit['b1'].value)
    b2 = str(edit['b2'].value)
    a1 = str(edit['a1'].value)
    a2 = str(edit['a2'].value)
    if cdata1[0] <= a1[5:] and '1' in b1:
        await channel.send(myid)
        await channel.send(embed=embed)
        edit['b1'].value = 'nmp'
        xl.save('db.xlsx')
    elif cdata1[0] >= a2[5:] and '1' in b2:
        await channel.send(myid)
        await channel.send(embed=embed)
        edit['b2'].value = 'nmp'
        xl.save('db.xlsx')


@client.command()
async def rs(ctx):
    xl = load_workbook('db.xlsx')
    edit = xl['Sheet1']
    edit['b1'].value = '1'
    edit['b2'].value = '1'
    a1 = str(edit['a1'].value)
    a2 = str(edit['a2'].value)
    xl.save('db.xlsx')
    await ctx.send('ok '+ a1 + a2)

@client.command()
async def nmin(ctx):
    xl = load_workbook('db.xlsx')
    edit = xl['Sheet1']
    edit['a1'].value = ctx.message.content
    a1 = str(edit['a1'].value)
    edit['b1'].value = '1'
    xl.save('db.xlsx')
    await ctx.send('nmin ok ' + a1)

@client.command()
async def nmax(ctx):
    xl = load_workbook('db.xlsx')
    edit = xl['Sheet1']
    edit['a2'].value = ctx.message.content
    a2 = str(edit['a2'].value)
    edit['b2'].value = '1'
    xl.save('db.xlsx')
    await ctx.send('nmax ok ' + a2)

@client.command()
async def fear(ctx):
    today = date.today()
    y = today.strftime("%Y")
    m = today.strftime("%m")
    d = today.strftime("%d")
    await ctx.send('https://alternative.me/images/fng/crypto-fear-and-greed-index-' + y + '-' + m[1:] + '-' + d + '.png')



client.run(TOKEN)

# # dcb > v1.3 racso574
